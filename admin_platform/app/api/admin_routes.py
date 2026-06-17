from fastapi import APIRouter, Request, Depends, Query, HTTPException, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from typing import Optional
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer
from PIL import Image as PILImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import sqlite3
import urllib.request
from urllib.parse import quote
import tempfile
import re
import os
import glob
import json
from PyPDF2 import PdfMerger
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors
from openpyxl.styles.colors import Color
from openpyxl.writer.excel import save_workbook
from sqlalchemy.orm import Session
from sqlalchemy import text

from ..db.database import get_db
from ..models.schemas import User
from ..services.auth_service import require_role, get_current_user

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

pdfmetrics.registerFont(TTFont('AppleGothic', '/System/Library/Fonts/Supplemental/AppleGothic.ttf'))

@router.get("/settings")
def settings_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver"], use_cache=False))
):
    settings_result = db.execute(text("SELECT value, start_date, end_date FROM settings WHERE key = 'max_overtime_hours'")).fetchone()
    max_overtime_hours = settings_result[0] if settings_result else 12
    start_date = settings_result[1] if settings_result else None
    end_date = settings_result[2] if settings_result else None

    response_data = {
        "request": request,
        "current_user": current_user,
        "max_overtime_hours": max_overtime_hours,
        "start_date": start_date,
        "end_date": end_date
    }
    
    # admin 권한일 때만 문서 승인자 설정 데이터 추가
    if current_user.role == "Admin":
        pdf_approver_result = db.execute(text("SELECT value FROM settings WHERE key = 'pdf_approver'")).fetchone()
        pdf_approver = pdf_approver_result[0] if pdf_approver_result else None
        employees = db.execute(text("SELECT name FROM employees WHERE role IN ('admin', 'manager', 'lead')")).fetchall()
        all_employees = db.execute(text("SELECT name FROM employees")).fetchall()
        document_manager_result = db.execute(text("SELECT value FROM settings WHERE key = 'document_manager'")).fetchone()
        document_manager = document_manager_result[0] if document_manager_result else None
        response_data["all_employees"] = all_employees
        response_data["document_manager"] = document_manager
        
        response_data["pdf_approver"] = pdf_approver
        response_data["employees"] = employees

    return render_template("settings.html", response_data)

@router.post("/settings")
def update_settings(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver"], use_cache=False)),
    max_overtime_hours: int = Form(...),
    start_date: str = Form(None),
    end_date: str = Form(None),
    pdf_approver: str = Form(None),
    document_manager: str = Form(None)
):
    # 시간외 근무 설정 업데이트 (모든 권한)
    db.execute(text("UPDATE settings SET value = :value, start_date = :start_date, end_date = :end_date WHERE key = 'max_overtime_hours'"), {"value": max_overtime_hours, "start_date": start_date, "end_date": end_date})
    
    # admin 권한일 때만 문서 승인자 설정 업데이트
    if current_user.role == "Admin":
        approver_setting = db.execute(text("SELECT 1 FROM settings WHERE key = 'pdf_approver'")).fetchone()
        if approver_setting:
            db.execute(text("UPDATE settings SET value = :value WHERE key = 'pdf_approver'"), {"value": pdf_approver})
        else:
            db.execute(text("INSERT INTO settings (key, value) VALUES ('pdf_approver', :value)"), {"value": pdf_approver})

        # 문서 관리자 설정 업데이트
        manager_setting = db.execute(text("SELECT 1 FROM settings WHERE key = 'document_manager'")).fetchone()
        if manager_setting:
            db.execute(text("UPDATE settings SET value = :value WHERE key = 'document_manager'"), {"value": document_manager})
        else:
            db.execute(text("INSERT INTO settings (key, value) VALUES ('document_manager', :value)"), {"value": document_manager})

        # 이전 manager 권한 제거 및 새 manager 권한 부여
        db.execute(text("UPDATE employees SET role = 'employee' WHERE role = 'manager'"))
        if document_manager:
            db.execute(text("UPDATE employees SET role = 'manager' WHERE name = :name"), {"name": document_manager})

    db.commit()
    return RedirectResponse(url="/settings", status_code=303)

@router.get("/stats")
def stats_page(
    request: Request,
    current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))
):
    return render_template("stats.html", {"request": request, "current_user": current_user})

@router.get("/api/stats/employee-status")
def get_employee_status(db: Session = Depends(get_db), current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))):
    employees_result = db.execute(text("SELECT name FROM employees WHERE name != 'admin'")).fetchall()
    employees = [row[0] for row in employees_result]

    status_data = []
    for emp_name in employees:
        # Calculate remaining compensatory hours
        overtime_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '시간외 근무' AND (status LIKE '%대기' OR status = '재신청' OR status = 'approved')"), {"name": emp_name}).fetchall()
        total_overtime_hours = 0
        for req in overtime_requests:
            content = json.loads(req[0])
            total_overtime_hours += content.get('calculated_compensatory_hours', 0)

        leave_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type IN ('대휴 사용', '대휴신청') AND (status LIKE '%대기' OR status = '재신청' OR status = 'approved')"), {"name": emp_name}).fetchall()
        used_leave_hours = 0
        for req in leave_requests:
            content = json.loads(req[0])
            used_leave_hours += content.get('hours', 0)

        remaining_hours = total_overtime_hours - used_leave_hours

        # Calculate remaining development cost
        dev_cost_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '자기개발비' AND (status LIKE '%대기' OR status = '재신청' OR status = 'approved')"), {"name": emp_name}).fetchall()
        used_dev_cost = 0
        for req in dev_cost_requests:
            content = json.loads(req[0])
            used_dev_cost += int(content.get('cost', '0'))
        
        remaining_dev_cost = 2000000 - used_dev_cost

        status_data.append({
            "name": emp_name,
            "remaining_compensatory_hours": remaining_hours,
            "remaining_dev_cost": remaining_dev_cost
        })

    return JSONResponse(content=status_data)

@router.get("/api/stats/overtime-hours")
def get_overtime_hours(
    period: str = Query("monthly", enum=["monthly", "yearly"]),
    month: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))
):
    employees_result = db.execute(text("SELECT name FROM employees WHERE name NOT IN ('admin', 'lead')")).fetchall()
    employees = [row[0] for row in employees_result]

    labels = []
    data = []
    today = datetime.today()

    for emp_name in employees:
        labels.append(emp_name)
        total_hours = 0

        if period == "monthly":
            if month:
                target_month_dt = datetime.strptime(month, "%Y-%m")
            else:
                target_month_dt = today
            
            end_date = (target_month_dt.replace(day=15)).strftime("%Y-%m-%d")
            start_date = (target_month_dt.replace(day=16) - relativedelta(months=1)).strftime("%Y-%m-%d")

        else: # yearly
            start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")
            end_date = today.replace(month=12, day=31).strftime("%Y-%m-%d")

        requests_result = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '시간외 근무' AND (status LIKE '%대기' OR status = '재신청' OR status = 'approved') AND json_extract(content, '$.work_date') BETWEEN :start_date AND :end_date"), {"name": emp_name, "start_date": start_date, "end_date": end_date}).fetchall()
        for req in requests_result:
            content = json.loads(req[0])
            total_hours += content.get('work_hours_weekday', 0) + content.get('work_hours_holiday', 0)
        
        data.append(total_hours)

    return JSONResponse(content={"labels": labels, "data": data})

@router.get("/admin-dashboard")
def admin_dashboard(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_name: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False)),
    overtime_page: int = 1,
    compensatory_page: int = 1,
    trip_page: int = 1,
    dev_page: int = 1,
    per_page: int = Query(10)
):
    # Fetch unique names for the dropdown
    names_result = db.execute(text("SELECT DISTINCT name FROM requests ORDER BY name ASC")).fetchall()
    names = [row[0] for row in names_result]

    # Date filtering logic
    if not start_date or not end_date:
        today = datetime.today()
        if today.day < 16:
            start_of_month = today.replace(day=16) - relativedelta(months=1)
        else:
            start_of_month = today.replace(day=16)
        end_of_month = start_of_month + relativedelta(months=1) - relativedelta(days=1)
        
        start_date = start_of_month.strftime("%Y-%m-%d")
        end_date = end_of_month.strftime("%Y-%m-%d")

    # Base query and params
    base_query = "SELECT * FROM requests WHERE created BETWEEN :start_date AND :end_date AND (status LIKE '%대기' OR status = '재신청' OR status = 'approved')"
    params = {"start_date": start_date, "end_date": end_date}

    if selected_name and selected_name != "all":
        base_query += " AND name = :name"
        params["name"] = selected_name

    # Pagination
    overtime_offset = (overtime_page - 1) * per_page
    compensatory_offset = (compensatory_page - 1) * per_page
    trip_offset = (trip_page - 1) * per_page
    dev_offset = (dev_page - 1) * per_page

    # Initialize request lists
    overtime_requests = []
    trip_requests = []
    self_dev_rows = []
    compensatory_leave_requests = []
    total_compensatory = 0
    total_trip = 0
    total_dev = 0
    total_overtime = 0

    if not request_type or request_type == "all" or request_type.startswith("시간외 근무"):
        overtime_query = f"{base_query} AND type = '시간외 근무'"
        if request_type == "시간외 근무 - 수당지급":
            overtime_query += " AND json_extract(content, '$.compensation') = '수당지급'"
        elif request_type == "시간외 근무 - 대체휴가":
            overtime_query += " AND json_extract(content, '$.compensation') = '대체휴가'"
        overtime_query += " ORDER BY id DESC LIMIT :limit OFFSET :offset"
        overtime_requests_raw = db.execute(text(overtime_query), {**params, "limit": per_page, "offset": overtime_offset}).fetchall()
        
        overtime_requests = []
        for r in overtime_requests_raw:
            content = json.loads(r[3])
            r = list(r)
            r.append(content.get('compensation'))
            overtime_requests.append(r)

        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '시간외 근무'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"

        if request_type == "시간외 근무 - 수당지급":
            count_query += " AND json_extract(content, '$.compensation') = '수당지급'"
        elif request_type == "시간외 근무 - 대체휴가":
            count_query += " AND json_extract(content, '$.compensation') = '대체휴가'"
        total_overtime = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type in ["대휴사용", "대휴신청"]:
        compensatory_leave_query = f"{base_query} AND type IN ('대휴 사용', '대휴신청') ORDER BY id DESC LIMIT :limit OFFSET :offset"
        compensatory_leave_requests = db.execute(text(compensatory_leave_query), {**params, "limit": per_page, "offset": compensatory_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type IN ('대휴 사용', '대휴신청')"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_compensatory = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type == "출장":
        trip_query = f"{base_query} AND type = '출장' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        trip_requests = db.execute(text(trip_query), {**params, "limit": per_page, "offset": trip_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '출장'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_trip = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type == "자기개발비":
        dev_query = f"{base_query} AND type = '자기개발비' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        self_dev_rows = db.execute(text(dev_query), {**params, "limit": per_page, "offset": dev_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '자기개발비'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_dev = db.execute(text(count_query), params).scalar_one()

    overtime_total_pages = (total_overtime + per_page - 1) // per_page
    compensatory_total_pages = (total_compensatory + per_page - 1) // per_page
    trip_total_pages = (total_trip + per_page - 1) // per_page
    dev_total_pages = (total_dev + per_page - 1) // per_page

    pending_count = db.execute(text("SELECT count(*) FROM requests WHERE status LIKE '%대기' OR status = '재신청' OR status = 'approved'")).scalar_one()
    approved_count = db.execute(text("SELECT count(*) FROM requests WHERE status LIKE '%대기' OR status = '재신청' OR status = 'approved'")).scalar_one()
    rejected_count = db.execute(text("SELECT count(*) FROM requests WHERE status = 'rejected'")).scalar_one()

    return render_template(
        "admin_dashboard.html",
        {
            "request": request,
            "overtime_requests": overtime_requests,
            "compensatory_leave_requests": compensatory_leave_requests,
            "trip_requests": trip_requests,
            "dev_requests": self_dev_rows,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "current_user": current_user,
            "start_date": start_date,
            "end_date": end_date,
            "names": names,
            "selected_name": selected_name,
            "request_type": request_type,
            "now": datetime.now(),
            "overtime_current_page": overtime_page,
            "overtime_total_pages": overtime_total_pages,
            "compensatory_current_page": compensatory_page,
            "compensatory_total_pages": compensatory_total_pages,
            "trip_current_page": trip_page,
            "trip_total_pages": trip_total_pages,
            "dev_current_page": dev_page,
            "dev_total_pages": dev_total_pages
        }
    )

@router.get("/admin-dashboard/pdf/merge")
def merge_pdfs(ids: str, db: Session = Depends(get_db), current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))):
    request_ids = [int(id) for id in ids.split(',')]

    request_ids_str = ",".join(map(str, request_ids))
    query = f"""SELECT r.id 
                 FROM requests r JOIN employees e ON r.name = e.name 
                 WHERE r.id IN ({request_ids_str}) 
                 ORDER BY e.emp_no ASC, json_extract(r.content, '$.work_date') ASC"""
    sorted_requests_result = db.execute(text(query)).fetchall()
    sorted_request_ids = [row[0] for row in sorted_requests_result]

    merger = PdfMerger()
    
    pdf_contents = []
    for request_id in sorted_request_ids:
        pdf_content = download_pdf(request_id, db, current_user)
        if pdf_content:
            pdf_contents.append(BytesIO(pdf_content))

    if not pdf_contents:
        return HTMLResponse(content="<script>alert('선택된 항목 중 PDF로 생성할 문서가 없습니다.'); window.history.back();</script>")

    for pdf_content in pdf_contents:
        merger.append(pdf_content)

    output_buffer = BytesIO()
    merger.write(output_buffer)
    merger.close()

    output_buffer.seek(0)
    return HTMLResponse(content=output_buffer.read(), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=merged-report.pdf"})



@router.get("/admin-dashboard/excel")
def export_to_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_name: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))
):
    if request_type:
        request_type = request_type.strip()
    
    if request_type != '시간외 근무 - 수당지급':
        return HTMLResponse(content="<script>alert('시간외 근무 - 수당지급을 선택하고 엑셀로 내보내기를 해주세요.'); window.history.back();</script>")

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "시간외 및 휴일근무 내역"
    
    # Get the current year and month for the title
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        title_date = start_date_obj.strftime('%Y년 %m월')
    else:
        title_date = datetime.now().strftime('%Y년 %m월')
    
    # Add title row with merged cells
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=f"{title_date} 시간외 및 휴일근무 내역")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Malgun Gothic', size=18, bold=True, color='FF000000')
    
    # Set row height for title row
    ws.row_dimensions[1].height = 36.75
    
    # Add empty row
    ws.append([])
    
    # Set headers at row 3
    headers = ["번호", "성명", "직번", "직급", "시간외 근로유형", "근무일", "요일", "근무시간 from", "근무시간 to", "부서장 인정시간", "연장근무 세부내역"]
    ws.append(headers)
    
    # Add another row for headers and merge cells except H and I
    subheaders = ["", "", "", "", "", "", "", "시작", "종료", "", ""]
    ws.append(subheaders)
    
    # Merge cells for headers (row 3 and 4) except columns H and I
    for col in [1, 2, 3, 4, 5, 6, 7, 10, 11]:  # A, B, C, D, E, F, G, J, K
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
    
    # Set specific text for H3 and I3
    ws.cell(row=3, column=8).value = "근무시간"
    ws.cell(row=3, column=9).value = "근무시간"
    
    # Merge H3 and I3 horizontally
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)

    # 헤더 스타일 정의
    # header_fill = PatternFill(start_color='FFffffff', end_color='FFffffff', fill_type='solid')
    header_fill = PatternFill(start_color='FF081F5C', end_color='FF081F5C', fill_type='solid')
    header_font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFFFF')
    
    # 헤더 셀에 스타일 적용
    for row in range(3, 5):  # Rows 3 and 4
        for col in range(1, 12):  # Columns A through K
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font

    # Fetch data
    where_clauses = []
    params = {}

    if start_date and end_date:
        where_clauses.append("r.created BETWEEN :start_date AND :end_date")
        params["start_date"] = start_date
        params["end_date"] = end_date

    where_clauses.append("(r.status LIKE '%대기' OR status = '재신청' OR status = 'approved')")
    where_clauses.append("r.type = '시간외 근무'")
    where_clauses.append("json_extract(r.content, '$.compensation') = '수당지급'")

    if selected_name and selected_name != "all":
        where_clauses.append("r.name = :name")
        params["name"] = selected_name

    where_clause = " AND ".join(where_clauses)
    query = f"""SELECT r.*, e.emp_no, e.position 
                 FROM requests r JOIN employees e ON r.name = e.name 
                 WHERE {where_clause} ORDER BY e.emp_no ASC, json_extract(r.content, '$.work_date') ASC"""

    requests_result = db.execute(text(query), params).fetchall()

    # Group requests by name
    from itertools import groupby
    def get_name(row):
        return row._mapping['name']

    grouped_requests = {k: list(v) for k, v in groupby(sorted(requests_result, key=get_name), key=get_name)}

    row_num = 5  # Start data from row 5 (after title row, empty row, header row, and subheader row)
    unique_id = 1
    for name, requests in grouped_requests.items():
        total_hours = 0
        start_row = row_num  # Save the starting row for this name
        for i, row_mapping in enumerate(requests):
            row = row_mapping._mapping
            content = json.loads(row['content'])
            work_date_str = content.get('work_date')
            
            day_of_week_korean = ''
            formatted_date = ''
            display_work_type = ''
            if work_date_str:
                try:
                    work_date = datetime.strptime(work_date_str, '%Y-%m-%d')
                    formatted_date = work_date.strftime('%Y.%m.%d')
                    day_map = {"Mon": "월", "Tue": "화", "Wed": "수", "Thu": "목", "Fri": "금", "Sat": "토", "Sun": "일"}
                    day_of_week_korean = day_map.get(work_date.strftime('%a'), '')

                    # 근무 유형을 실제 요일 기준으로 판단
                    if work_date.weekday() < 5: # 0:월, 1:화, 2:수, 3:목, 4:금
                        display_work_type = "평일연장근로(최대3시간)"
                    else: # 5:토, 6:일
                        display_work_type = "휴일근로(최대6시간)"
                except ValueError:
                    pass

            work_time_range = content.get('work_time_range', '-').split('-')
            
            if i == 0:
                ws.cell(row=row_num, column=1, value=unique_id)
                unique_id += 1
            else:
                ws.cell(row=row_num, column=1, value="")

            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['emp_no'])
            ws.cell(row=row_num, column=4, value=row['position'])
            ws.cell(row=row_num, column=5, value=display_work_type)
            ws.cell(row=row_num, column=6, value=formatted_date)
            ws.cell(row=row_num, column=7, value=day_of_week_korean)
            ws.cell(row=row_num, column=8, value=work_time_range[0])
            ws.cell(row=row_num, column=9, value=work_time_range[1] if len(work_time_range) > 1 else '')
            hours = content.get('work_hours_weekday', 0) + content.get('work_hours_holiday', 0)
            ws.cell(row=row_num, column=10, value=hours)
            
            # J열(10번 열)에 시간 값 설정 및 자동 줄바꿈 서식 적용
            j_cell = ws.cell(row=row_num, column=10)
            j_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            total_hours += hours
            ws.cell(row=row_num, column=11, value=content.get('reason_detail', ''))
            row_num += 1

        # Merge cells in column A, B, C, D (no, name, emp_no, position) if there are multiple rows for this name
        if row_num > start_row + 1:  # If there's more than one row for this name
            # A열(번호) 병합
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
            # B열(성명) 병합
            ws.merge_cells(start_row=start_row, start_column=2, end_row=row_num-1, end_column=2)
            # C열(직번) 병합
            ws.merge_cells(start_row=start_row, start_column=3, end_row=row_num-1, end_column=3)
            # D열(직급) 병합
            ws.merge_cells(start_row=start_row, start_column=4, end_row=row_num-1, end_column=4)
        
        # Add summary row
        ws.cell(row=row_num, column=10, value=total_hours)
        
        # 합계 행 설정 (A~I 셀 병합, "계" 표시, 배경색 설정)
        # A부터 I까지 셀 병합
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        merged_cell = ws.cell(row=row_num, column=1)
        merged_cell.value = "계"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 배경색 설정 (16진수 D9D9D9)
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # 병합된 셀과 J열, K열 셀에 배경색 적용
        for col in range(1, 12):  # A부터 K열까지
            cell = ws.cell(row=row_num, column=col)
            cell.fill = gray_fill
        

        # Add summary row
        ws.cell(row=row_num, column=10, value=total_hours)
        row_num += 1

    # Set font and alignment for all cells
    default_font = Font(name='Malgun Gothic', size=9)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 열 너비 설정
    ws.column_dimensions['A'].width = 5.17   # A열
    ws.column_dimensions['B'].width = 11.33  # B열
    ws.column_dimensions['C'].width = 8      # C열
    ws.column_dimensions['D'].width = 11.83  # D열
    ws.column_dimensions['E'].width = 19.83  # E열
    ws.column_dimensions['F'].width = 8.83   # F열
    ws.column_dimensions['K'].width = 47     # K열
    
    header_rows = {3, 4}

    for row_idx, row in enumerate(ws.iter_rows(), 1):  # row_idx starts from 1
        for cell in row:
            # Skip title row (row 1) to preserve its font settings
            if row_idx != 1 and row_idx not in header_rows:  # Apply default font only to non-header rows
                cell.font = default_font
            if cell.column != 11:  # "연장근무 세부내역" column
                cell.alignment = center_alignment
            # k3, k4 셀(K열의 3행과 4행)에도 수직 가운데 정렬 적용
            elif row_idx in [3, 4] and cell.column == 11:
                cell.alignment = center_alignment

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "시간외_및_휴일근무_내역.xlsx"

    return HTMLResponse(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('시간외_및_휴일근무_내역.xlsx')}"}
    )
# @router.get("/admin-dashboard/excel")
# def export_to_excel(
#     start_date: Optional[str] = Query(None),
#     end_date: Optional[str] = Query(None),
#     selected_name: Optional[str] = Query(None),
#     request_type: Optional[str] = Query(None),
#     db: Session = Depends(get_db),
#     current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))
# ):
#     # Date filtering logic
#     if not start_date or not end_date:
#         today = datetime.today()
#         if today.day < 16:
#             start_of_month = today.replace(day=16) - relativedelta(months=1)
#         else:
#             start_of_month = today.replace(day=16)
#         end_of_month = start_of_month + relativedelta(months=1) - relativedelta(days=1)
        
#         start_date = start_of_month.strftime("%Y-%m-%d")
#         end_date = end_of_month.strftime("%Y-%m-%d")

#     # Base query and params
#     base_query = "SELECT r.*, e.emp_no, e.dept, e.position FROM requests r JOIN employees e ON r.name = e.name WHERE r.created BETWEEN :start_date AND :end_date"
#     params = {"start_date": start_date, "end_date": end_date}

#     if selected_name and selected_name != "all":
#         base_query += " AND r.name = :name"
#         params["name"] = selected_name

#     if request_type and request_type != "all":
#         if request_type == "시간외 근무 - 수당지급":
#             base_query += " AND r.type = '시간외 근무' AND json_extract(r.content, '$.compensation') = '수당지급'"
#         elif request_type == "시간외 근무 - 대체휴가":
#             base_query += " AND r.type = '시간외 근무' AND json_extract(r.content, '$.compensation') = '대체휴가'"
#         else:
#             base_query += " AND r.type = :type"
#             params["type"] = request_type

#     base_query += " ORDER BY e.emp_no ASC, r.created DESC"
    
#     requests_result = db.execute(text(base_query), params).fetchall()
    
#     # Create a new workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "근무 내역"
    
#     # Add headers
#     headers = ["이름", "사번", "부서", "직급", "신청 유형", "신청일", "상태"]
    
#     if request_type and request_type.startswith("시간외 근무"):
#         headers.extend(["근무일", "근무 유형", "평일 시간", "휴일 시간", "보상 유형", "근무 장소", "근무 시간대", "신청 사유", "상세 사유"])
#     elif request_type == "출장":
#         headers.extend(["시작일", "종료일", "출장 지역", "출장 기관", "이동 수단", "목적"])
#     elif request_type == "자기개발비":
#         headers.extend(["수강 항목", "목적", "수강 내용", "비용", "참고 사이트"])
#     elif request_type in ["대휴사용", "대휴신청"]:
#         headers.extend(["사용일", "사용 시간"])
    
#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')
#         cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
#     # Add data
#     for row_num, req in enumerate(requests_result, 2):
#         req_dict = dict(req._mapping)
#         content = json.loads(req_dict['content']) if req_dict['content'] else {}
        
#         # Common fields
#         ws.cell(row=row_num, column=1).value = req_dict['name']
#         ws.cell(row=row_num, column=2).value = req_dict['emp_no']
#         ws.cell(row=row_num, column=3).value = req_dict['dept']
#         ws.cell(row=row_num, column=4).value = req_dict['position']
#         ws.cell(row=row_num, column=5).value = req_dict['type']
#         ws.cell(row=row_num, column=6).value = req_dict['created'].split(' ')[0] if req_dict['created'] else ''
#         ws.cell(row=row_num, column=7).value = req_dict['status']
        
#         # Type-specific fields
#         if req_dict['type'] == '시간외 근무':
#             ws.cell(row=row_num, column=8).value = content.get('work_date', '')
#             ws.cell(row=row_num, column=9).value = content.get('work_type', '')
#             ws.cell(row=row_num, column=10).value = content.get('work_hours_weekday', 0)
#             ws.cell(row=row_num, column=11).value = content.get('work_hours_holiday', 0)
#             ws.cell(row=row_num, column=12).value = content.get('compensation', '')
#             ws.cell(row=row_num, column=13).value = content.get('work_location', '')
#             ws.cell(row=row_num, column=14).value = content.get('work_time_range', '')
#             ws.cell(row=row_num, column=15).value = content.get('reason_type', '')
#             ws.cell(row=row_num, column=16).value = content.get('reason_detail', '')
#         elif req_dict['type'] == '출장':
#             ws.cell(row=row_num, column=8).value = content.get('start_date', '')
#             ws.cell(row=row_num, column=9).value = content.get('end_date', '')
#             ws.cell(row=row_num, column=10).value = content.get('region', '') or content.get('region_other', '')
#             ws.cell(row=row_num, column=11).value = content.get('organization', '')
#             ws.cell(row=row_num, column=12).value = content.get('transport', '')
#             ws.cell(row=row_num, column=13).value = content.get('purpose', '')
#         elif req_dict['type'] == '자기개발비':
#             ws.cell(row=row_num, column=8).value = content.get('course_title', '')
#             ws.cell(row=row_num, column=9).value = content.get('purpose', '')
#             ws.cell(row=row_num, column=10).value = content.get('course_content', '')
#             ws.cell(row=row_num, column=11).value = content.get('cost', 0)
#             ws.cell(row=row_num, column=12).value = content.get('reference_site', '')
#         elif req_dict['type'] in ['대휴 사용', '대휴신청']:
#             ws.cell(row=row_num, column=8).value = content.get('leave_date', '')
#             ws.cell(row=row_num, column=9).value = content.get('hours', 0)
    
#     # Auto-adjust column widths
#     for column in ws.columns:
#         max_length = 0
#         column_letter = column[0].column_letter
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width
    
#     # Save to a BytesIO object
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
    
#     # Return the Excel file
#     filename = f"근무내역_{start_date}_{end_date}.xlsx"
#     headers = {
#         'Content-Disposition': f'attachment; filename="{quote(filename)}"'
#     }
    
#     return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

def download_pdf(request_id: int, db: Session, current_user: User) -> bytes:
    try:
        print(f"--- Generating PDF for request_id: {request_id} ---")
        request_data_result = db.execute(text("SELECT * FROM requests WHERE id = :id"), {"id": request_id}).fetchone()

        if not request_data_result:
            print("!!! Request not found")
            raise HTTPException(status_code=404, detail="Request not found")

        request_data = dict(request_data_result._mapping)
        print(f"Request Data: {request_data}")

        request_content_str = request_data.get('content', '{}')
        request_content = json.loads(request_content_str)
        print(f"Request Content: {request_content}")


        # Determine the approver for the PDF
        pdf_approver_setting = db.execute(text("SELECT value FROM settings WHERE key = 'pdf_approver'")).fetchone()
        
        approver_name = None
        if pdf_approver_setting and pdf_approver_setting[0]:
            approver_name = pdf_approver_setting[0]
        else:
            approver_name = request_data.get('approver')

        if not approver_name:
            approver_name = "관리자"
            print("!!! No approver found, using default")

        signature_data = None
        approver_position = "(미지정)"

        if approver_name:
            approver_info_result = db.execute(text("SELECT signature, position FROM employees WHERE name = :name"), {"name": approver_name}).fetchone()
            if approver_info_result:
                approver_info_mapping = dict(approver_info_result._mapping)
                print(f"Approver Info: {approver_info_mapping}")
                signature_data = approver_info_mapping.get('signature')
                approver_position = approver_info_mapping.get('position', '(미지정)')

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        styles['h1'].fontName = 'AppleGothic'
        styles['h1'].alignment = TA_CENTER
        styles['Normal'].fontName = 'AppleGothic'
        styles['Normal'].fontSize = 12

        title = request_data.get('type', '문서')
        story.append(Paragraph(f"{title} 승인 확인서", styles['h1']))
        story.append(Spacer(1, 24))

        request_data_list = [[Paragraph(f"<b>신청자:</b> {request_data['name']}", styles['Normal'])], [Paragraph(f"<b>신청일:</b> {request_data['created'].split(' ')[0]}", styles['Normal'])], [Paragraph("", styles['Normal'])], [Paragraph("", styles['Normal'])]]
        key_map = {
            "work_type": "근무 유형", "work_date": "근무일", "work_hours_weekday": "평일 근무시간",
            "work_hours_holiday": "휴일 근무시간", "leave_date": "대휴 사용일", "reason_type": "신청 사유", "reason_detail": "상세 사유",
            "work_location": "근무 장소", "compensation": "보상 유형", "course_title": "수강 항목",
            "purpose": "목적", "purpose_other": "목적 (기타)", "course_content": "수강 내용", "cost": "비용",
            "start_date": "시작일", "end_date": "종료일", "reference_site": "참고 사이트",
            "region": "출장 지역", "region_other": "출장 지역 (기타)", "organization": "출장 기관",
            "transport": "이동 수단", "hours": "사용 시간"
        }
        
        for key, value in request_content.items():
            if key == "work_time_range":
                continue
            if key in key_map and value:  # None이나 빈 문자열이 아닌 경우에만 추가
                # 기타 선택 시 중복 표시 방지
                if key == "region" and value == "기타" and request_content.get("region_other"):
                    continue
                if key == "purpose" and value == "기타" and request_content.get("purpose_other"):
                    continue
                # 시간외근무 시간 정보에 시간 범위 추가
                if key in ["work_hours_weekday", "work_hours_holiday"]:
                    work_time_range = request_content.get("work_time_range", "")
                    if work_time_range:
                        display_value = f"{work_time_range} {value}hr"
                    else:
                        display_value = f"{value}hr"
                else:
                    display_value = str(value)
                request_data_list.append([Paragraph(f"<b>{key_map[key]}:</b> {display_value}", styles['Normal'])])

        request_table = Table(request_data_list, hAlign='LEFT')
        request_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'AppleGothic'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        story.append(request_table)
        story.append(Spacer(1, 48))

        story.append(Paragraph("상기 신청을 승인합니다.", styles['Normal']))
        story.append(Spacer(1, 24))

        story.append(Paragraph(f"<b>승인 날짜:</b> {request_data['created'].split(' ')[0]}", styles['Normal']))
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>승인자:</b> {approver_position} {approver_name}", styles['Normal']))

        if signature_data:
            print("Signature data found, attempting to process.")
            try:
                img_file = BytesIO(signature_data)
                pil_img = PILImage.open(img_file)
                pil_img.verify()
                img_file.seek(0)
                story.append(Image(img_file, width=50, height=50, hAlign='LEFT'))
                print("Successfully added signature image to story.")
            except Exception as e:
                print(f"!!! Could not process signature image: {e}")
        else:
            print("!!! No signature data found for approver.")

        if not story:
            print("!!! Story is empty, cannot build PDF.")
            raise HTTPException(status_code=500, detail="PDF 생성에 실패했습니다: 내용이 없습니다.")

        print(f"Final story length: {len(story)}")
        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    except Exception as e:
        print(f"!!! Error generating PDF: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF 생성 중 오류가 발생했습니다: {str(e)}")

@router.get("/admin-dashboard/pdf/{request_id}")
def download_pdf_route(request_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    request_owner_result = db.execute(text("SELECT name FROM requests WHERE id = :id"), {"id": request_id}).fetchone()
    request_owner = request_owner_result[0] if request_owner_result else None

    if not request_owner or (current_user.name != request_owner and current_user.role not in ["Admin", "Approver", "Lead", "Manager"]):
        raise HTTPException(status_code=403, detail="이 PDF에 접근할 권한이 없습니다.")
    
    pdf_content = download_pdf(request_id, db, current_user)
    if not pdf_content:
        raise HTTPException(status_code=404, detail="PDF를 생성할 수 없습니다.")
    
    buffer = BytesIO(pdf_content)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=document-{request_id}.pdf"}
    )

@router.get("/reject/cancel/{request_id}")
def cancel_rejection(request_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role(["Admin", "Approver"]))):
    request_info = db.execute(text("SELECT approver FROM requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not request_info:
        raise HTTPException(status_code=404, detail="Request not found")

    previous_status = f"{dict(request_info._mapping).get('approver', 'manager')} 승인 대기"

    db.execute(text("UPDATE requests SET status = :status, reject_reason = NULL WHERE id = :id"), {
        "status": previous_status,
        "id": request_id
    })
    db.commit()
    return RedirectResponse(url="/approve-list", status_code=303)

# 템플릿 렌더링 함수 (캐시 없음)
import os
from jinja2 import Environment, FileSystemLoader
import json
from datetime import datetime
from fastapi.responses import HTMLResponse

template_dir = os.path.join(os.path.dirname(__file__), "..", "templates")
jinja_env = Environment(
    loader=FileSystemLoader(template_dir),
    cache_size=0,
    auto_reload=True
)
jinja_env.filters['fromjson'] = json.loads
jinja_env.filters['to_datetime'] = lambda s: datetime.strptime(s, '%Y-%m-%d') if s else None

def render_template(template_name: str, context: dict):
    template = jinja_env.get_template(template_name)
    html_content = template.render(context)
    return HTMLResponse(content=html_content)

@router.get("/admin-dashboard/api/load-more")
def load_more_requests(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_name: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    table_type: str = Query(...),
    offset: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["Admin", "Approver", "Lead", "Manager"], use_cache=False))
):
    if not start_date or not end_date:
        today = datetime.today()
        if today.day < 16:
            start_of_month = today.replace(day=16) - relativedelta(months=1)
        else:
            start_of_month = today.replace(day=16)
        end_of_month = start_of_month + relativedelta(months=1) - relativedelta(days=1)
        start_date = start_of_month.strftime("%Y-%m-%d")
        end_date = end_of_month.strftime("%Y-%m-%d")

    params = {"start_date": start_date, "end_date": end_date}
    base_query = "SELECT id, name, created, content, status FROM requests WHERE created BETWEEN :start_date AND :end_date"
    
    if selected_name and selected_name != "all":
        base_query += " AND name = :name"
        params["name"] = selected_name

    if table_type == "overtime":
        query = f"{base_query} AND type = '시간외 근무'"
        if request_type == "시간외 근무 - 수당지급":
            query += " AND json_extract(content, '$.compensation') = '수당지급'"
        elif request_type == "시간외 근무 - 대체휴가":
            query += " AND json_extract(content, '$.compensation') = '대체휴가'"
        query += " ORDER BY id DESC LIMIT :limit OFFSET :offset"
        requests_raw = db.execute(text(query), {**params, "limit": limit, "offset": offset}).fetchall()
        requests = []
        for r in requests_raw:
            content = json.loads(r[3])
            r = list(r)
            r.append(content.get('compensation'))
            requests.append(r)
    elif table_type == "compensatory":
        query = f"{base_query} AND type IN ('대휴 사용', '대휴신청') ORDER BY id DESC LIMIT :limit OFFSET :offset"
        requests = db.execute(text(query), {**params, "limit": limit, "offset": offset}).fetchall()
    elif table_type == "trip":
        query = f"{base_query} AND type = '출장' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        requests = db.execute(text(query), {**params, "limit": limit, "offset": offset}).fetchall()
    elif table_type == "dev":
        query = f"{base_query} AND type = '자기개발비' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        requests = db.execute(text(query), {**params, "limit": limit, "offset": offset}).fetchall()
    else:
        return JSONResponse(content={"error": "Invalid table_type"}, status_code=400)

    result = []
    for req in requests:
        content = json.loads(req[3]) if req[3] else {}
        result.append({
            "id": req[0],
            "name": req[1],
            "created": req[2],
            "content": content,
            "status": req[4],
            "compensation": req[5] if len(req) > 5 else None
        })

    return JSONResponse(content={"requests": result, "has_more": len(result) == limit})

@router.get("/approve/cancel/{request_id}")
def cancel_approval(request_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role(["Admin", "Approver"]))):
    """승인 취소 - 승인된 건을 다시 대기 상태로 변경"""
    request_info = db.execute(text("SELECT approver, status FROM requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not request_info:
        raise HTTPException(status_code=404, detail="Request not found")
    
    request_mapping = dict(request_info._mapping)
    current_status = request_mapping.get('status', '')
    
    # 승인된 건만 취소 가능
    if 'approved' not in current_status.lower():
        raise HTTPException(status_code=400, detail="승인된 건만 취소할 수 있습니다.")
    
    # 이전 대기 상태로 복원
    approver = request_mapping.get('approver', 'manager')
    previous_status = f"{approver} 승인 대기"
    
    db.execute(text("UPDATE requests SET status = :status WHERE id = :id"), {
        "status": previous_status,
        "id": request_id
    })
    db.commit()
    
    return RedirectResponse(url="/approve-list", status_code=303)
